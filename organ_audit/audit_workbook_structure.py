from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path("organ_audit/output/post_2015_examination.xlsx")
OUT = Path("organ_audit/analysis/workbook_structure.json")


def color_value(color) -> str | None:
    if color is None:
        return None
    if getattr(color, "type", None) == "rgb":
        return color.rgb
    if getattr(color, "type", None) == "indexed":
        return f"indexed:{color.indexed}"
    if getattr(color, "type", None) == "theme":
        return f"theme:{color.theme}:tint:{color.tint}"
    return str(getattr(color, "value", None))


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False)
    result: dict[str, object] = {
        "sheetnames": wb.sheetnames,
        "properties": {
            "creator": wb.properties.creator,
            "last_modified_by": wb.properties.lastModifiedBy,
            "created": wb.properties.created.isoformat() if wb.properties.created else None,
            "modified": wb.properties.modified.isoformat() if wb.properties.modified else None,
            "title": wb.properties.title,
            "subject": wb.properties.subject,
            "description": wb.properties.description,
        },
        "defined_names": [str(item) for item in wb.defined_names.values()],
        "sheets": [],
    }

    for ws in wb.worksheets:
        fills: Counter[str] = Counter()
        fonts: Counter[str] = Counter()
        alignments: Counter[str] = Counter()
        number_formats: Counter[str] = Counter()
        style_ids: Counter[int] = Counter()
        comments: list[dict[str, object]] = []
        formulas: list[dict[str, object]] = []
        hyperlinks: list[dict[str, object]] = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                style_ids[cell.style_id] += 1
                fills[f"{cell.fill.fill_type}:{color_value(cell.fill.fgColor)}:{color_value(cell.fill.bgColor)}"] += 1
                fonts[
                    f"bold={cell.font.bold};italic={cell.font.italic};color={color_value(cell.font.color)}"
                ] += 1
                alignments[
                    f"h={cell.alignment.horizontal};v={cell.alignment.vertical};wrap={cell.alignment.wrap_text}"
                ] += 1
                number_formats[cell.number_format] += 1
                if cell.comment:
                    comments.append(
                        {
                            "cell": cell.coordinate,
                            "author": cell.comment.author,
                            "text": cell.comment.text,
                        }
                    )
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if cell.hyperlink:
                    hyperlinks.append(
                        {"cell": cell.coordinate, "target": cell.hyperlink.target}
                    )

        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_columns = [key for key, dim in ws.column_dimensions.items() if dim.hidden]
        sheet = {
            "title": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "auto_filter": ws.auto_filter.ref,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns,
            "data_validations": len(ws.data_validations.dataValidation),
            "conditional_formatting_rules": len(ws.conditional_formatting),
            "comments": comments,
            "formulas": formulas,
            "hyperlinks": hyperlinks,
            "style_ids": dict(style_ids),
            "fills": dict(fills),
            "fonts": dict(fonts),
            "alignments": dict(alignments),
            "number_formats": dict(number_formats),
        }
        result["sheets"].append(sheet)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
